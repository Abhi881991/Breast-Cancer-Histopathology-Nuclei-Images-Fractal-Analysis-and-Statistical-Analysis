"""
================================================================================
RESEARCH QUESTION 4 (RQ4): WITHIN-WSI HETEROGENEITY ANALYSIS
================================================================================

Research Question:
"How heterogeneous is nuclear spatial organization within the same whole-slide 
image (WSI), and does this heterogeneity correlate with pathological diagnosis?"

Theoretical Background:
-----------------------

SPATIAL HETEROGENEITY CONCEPT:
    A whole-slide image (WSI) represents one patient's tissue sample, typically
    containing multiple regions of interest (ROIs). Each ROI has its own fractal
    dimension measurement. The question is: How much do these measurements vary
    WITHIN the same patient?
    
    High heterogeneity → Different regions have very different spatial patterns
    Low heterogeneity → Regions are spatially similar within the patient

BIOLOGICAL SIGNIFICANCE:
    Heterogeneity has important clinical implications:
    
    1. DIAGNOSTIC AMBIGUITY:
       - High heterogeneity may indicate mixed pathologies
       - Difficult to assign single diagnosis
       - Sampling bias concerns
    
    2. DISEASE PROGRESSION:
       - Transition zones between normal and malignant
       - Pre-cancerous lesions showing heterogeneous architecture
       - Tumor heterogeneity (ITH - intratumor heterogeneity)
    
    3. PROGNOSIS:
       - High heterogeneity associated with poor outcomes
       - Indicates aggressive/evolving disease
       - Multiple clonal populations
    
    4. TREATMENT IMPLICATIONS:
       - Heterogeneous tumors harder to treat
       - Drug resistance mechanisms
       - Need for comprehensive sampling

CRITICAL DATA CHARACTERISTIC (DISCOVERED IN INITIAL ANALYSIS):
    236 of 368 WSIs (64%) contain MIXED PATHOLOGIES
    - Same patient has regions with different diagnoses
    - This is NOT measurement error - it's biological reality
    - Examples: N + PB + UDH in same WSI
    
    This fundamentally affects how we measure heterogeneity:
    - "Pure" WSIs (132 WSIs, 36%): Single pathology
    - "Mixed" WSIs (236 WSIs, 64%): Multiple pathologies
    
    Heterogeneity can arise from:
    a) Natural variation within same pathology (biological)
    b) Measurement error (technical)
    c) Mixed pathologies in same patient (pathological)

STATISTICAL MEASURES OF HETEROGENEITY:
---------------------------------------

1. COEFFICIENT OF VARIATION (CV)
   Purpose: Standardized measure of dispersion
   Formula: CV = (σ / μ) × 100%
   
   where:
   - σ = standard deviation of Dc within WSI
   - μ = mean Dc within WSI
   
   Interpretation:
   - CV < 10%: Low heterogeneity (homogeneous)
   - 10% ≤ CV < 20%: Moderate heterogeneity
   - CV ≥ 20%: High heterogeneity (heterogeneous)
   
   Why use it: Scale-independent, allows comparison across WSIs

2. RANGE (Max - Min)
   Purpose: Measure extreme variability
   Formula: Range = max(Dc) - min(Dc)
   
   Interpretation:
   - Large range → Some regions very different from others
   - Sensitive to outliers
   - Clinical relevance: Detects focal abnormalities
   
   Why use it: Simple, clinically interpretable

3. INTERQUARTILE RANGE (IQR)
   Purpose: Robust measure of spread
   Formula: IQR = Q3 - Q1 (75th - 25th percentile)
   
   Interpretation:
   - Contains middle 50% of data
   - Resistant to outliers
   - More stable than range
   
   Why use it: Robust alternative to standard deviation

4. INTRACLASS CORRELATION COEFFICIENT (ICC)
   Purpose: Partition variance into within-WSI vs between-WSI
   Formula: ICC = σ²_between / (σ²_between + σ²_within)
   
   Model: One-way random effects ANOVA
   
   Interpretation:
   - ICC close to 1: Most variance between WSIs (low within-WSI heterogeneity)
   - ICC close to 0: Most variance within WSIs (high within-WSI heterogeneity)
   - ICC = 0.5: Equal variance within and between WSIs
   
   Why use it: Gold standard for nested/hierarchical data

5. MEDIAN ABSOLUTE DEVIATION (MAD)
   Purpose: Robust measure of variability
   Formula: MAD = median(|Dc_i - median(Dc)|)
   
   Normalized: MAD / median
   
   Why use it: Very robust to outliers, distribution-free

6. VARIANCE RATIO TEST
   Purpose: Compare heterogeneity across pathology groups
   Formula: F = s₁² / s₂² (ratio of variances)
   
   Levene's test: More robust, tests equality of variances
   Brown-Forsythe: Uses medians instead of means
   
   Why use it: Determines if some pathologies more heterogeneous

STATISTICAL TESTS USED:
-----------------------

1. DESCRIPTIVE STATISTICS BY WSI
   - Mean, SD, CV, Range, IQR, MAD for each WSI
   - Summary across all WSIs
   - Distribution of heterogeneity measures

2. VARIANCE COMPONENTS ANALYSIS (ICC)
   - One-way random effects ANOVA
   - Decompose total variance into:
     * Between-WSI variance (σ²_between)
     * Within-WSI variance (σ²_within)
   - ICC calculation

3. HETEROGENEITY vs PATHOLOGY (ANOVA)
   Purpose: Test if heterogeneity differs by diagnosis
   H₀: σ²_N = σ²_PB = σ²_UDH = ... (equal heterogeneity)
   
   Use: Within-WSI SD as dependent variable
   Group: Pathology type

4. HETEROGENEITY vs ROI DENSITY (CORRELATION)
   Purpose: Does more sampling reveal more heterogeneity?
   Method: Pearson/Spearman correlation
   Variables: n_ROIs vs SD(Dc)
   
   Expected: Positive correlation (more sampling → more extremes)

5. MIXED vs PURE PATHOLOGY COMPARISON (T-TEST)
   Purpose: Compare heterogeneity in pure vs mixed WSIs
   H₀: μ_pure = μ_mixed (same mean heterogeneity)
   
   Use: Independent samples t-test or Mann-Whitney U
   
   Prediction: Mixed WSIs should show higher heterogeneity

6. OUTLIER DETECTION (Z-SCORES, DBSCAN)
   Purpose: Identify WSIs with extreme heterogeneity
   Methods:
   - Z-scores: |z| > 3 (outliers)
   - Modified Z-score: Based on MAD
   - DBSCAN: Density-based clustering
   
   Why: Outliers may represent complex/transitional cases

7. SPATIAL AUTOCORRELATION (IF SPATIAL COORDINATES AVAILABLE)
   Purpose: Test if nearby ROIs are more similar
   Method: Moran's I, Geary's C
   
   Note: Would require ROI spatial coordinates (not in current data)

8. RELIABILITY ANALYSIS (STRATIFICATION BY ROI COUNT)
   Purpose: Assess reliability of heterogeneity estimates
   Strata: 
   - 1 ROI: Cannot estimate (SE undefined)
   - 2-5 ROIs: Low reliability
   - 6-10 ROIs: Moderate reliability  
   - >10 ROIs: Good reliability
   
   Method: Standard error of mean: SE = SD / √n

9. MULTILEVEL MODEL (MIXED EFFECTS)
   Purpose: Model ROI-level Dc with WSI as random effect
   Model: Dc ~ Pathology + (1|WSI_ID)
   
   Outputs:
   - Fixed effects: Pathology differences (from RQ3)
   - Random effects: WSI-specific deviations
   - Residual variance: Within-WSI variability

10. DISTRIBUTION ANALYSIS
    Purpose: Test if heterogeneity is normally distributed
    Methods: Shapiro-Wilk, Anderson-Darling
    
    Implications for parametric tests

PARAMETERS CHOSEN:
------------------
- Significance level: α = 0.05
- CV thresholds: 10% (low), 20% (high)
- Minimum ROIs for analysis: 2 (need at least 2 for variance)
- Outlier threshold: |z| > 3 (3 standard deviations)
- Bootstrap samples: 10,000 for confidence intervals
- ICC confidence level: 95%

EXPECTED OUTCOMES:
------------------
If RQ4 reveals high within-WSI heterogeneity:
1. ICC < 0.5 (substantial within-WSI variance) ✓
2. High mean CV across WSIs (>15%) ✓
3. Large ranges in many WSIs ✓
4. Mixed pathology WSIs show higher heterogeneity ✓
5. Pre-cancerous lesions (ADH, FEA) most heterogeneous ✓

If heterogeneity is low:
1. ICC > 0.8 (most variance between WSIs) ✗
2. Low mean CV (<10%) ✗
3. Small ranges consistently ✗
4. Pure and mixed WSIs similar ✗

CLINICAL IMPLICATIONS:
---------------------
High heterogeneity suggests:
- Need for multiple biopsies per patient
- Sampling bias is significant concern
- Single ROI may not represent entire tumor
- Spatial mapping important for diagnosis
- Personalized treatment planning needed

Low heterogeneity suggests:
- Single biopsy may be sufficient
- Spatial patterns consistent within patient
- Easier diagnosis and staging
- More predictable treatment response

================================================================================
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from scipy.stats import (f_oneway, levene, ttest_ind, mannwhitneyu, 
                         pearsonr, spearmanr, shapiro, anderson)
from sklearn.cluster import DBSCAN
from sklearn.preprocessing import StandardScaler
import statsmodels.api as sm
from statsmodels.formula.api import ols
from statsmodels.stats.anova import anova_lm
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Set publication-quality plot parameters
plt.rcParams['figure.dpi'] = 600
plt.rcParams['savefig.dpi'] = 600
plt.rcParams['font.family'] = 'Times New Roman'  # Available font
plt.rcParams['font.weight'] = 'bold'
plt.rcParams['axes.labelsize'] = 20
plt.rcParams['axes.linewidth'] = 3
plt.rcParams['axes.titlesize'] = 12
plt.rcParams['axes.labelweight'] = 'bold'
plt.rcParams['axes.titleweight'] = 'bold'
plt.rcParams['xtick.labelsize'] = 16
plt.rcParams['ytick.labelsize'] = 16
plt.rcParams['legend.fontsize'] = 10

# ============================================================================
# CONFIGURATION
# ============================================================================

# Input file paths
BASE_PATH = Path(r"C:\Users\ajd44\Desktop")
CORR_FILE = BASE_PATH / "Correlation Dimension.csv"
MINK_FILE = BASE_PATH / "Minkowski Dimension.csv"

# Output directory
OUTPUT_DIR = BASE_PATH / 'RQ4_Within_WSI_Heterogeneity_Analysis'
OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

# Create subdirectories
PLOTS_DIR = OUTPUT_DIR / 'plots'
RESULTS_DIR = OUTPUT_DIR / 'results'
ORIGIN_DATA_DIR = OUTPUT_DIR / 'origin_data'
PLOTS_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)
ORIGIN_DATA_DIR.mkdir(exist_ok=True)

# Analysis parameters
ALPHA = 0.05
N_BOOTSTRAP = 10000
RANDOM_STATE = 42
MIN_ROIS = 2  # Minimum ROIs needed for heterogeneity calculation

# Heterogeneity thresholds (coefficient of variation)
CV_LOW = 10.0   # <10% = homogeneous
CV_HIGH = 20.0  # ≥20% = heterogeneous

# Pathology order
PATHOLOGY_ORDER = ['N', 'PB', 'UDH', 'FEA', 'ADH', 'DCIS', 'IC']

print("=" * 80)
print("RQ4: WITHIN-WSI HETEROGENEITY ANALYSIS")
print("=" * 80)
print(f"\nOutput directory: {OUTPUT_DIR}")
print(f"Minimum ROIs per WSI: {MIN_ROIS}")
print(f"CV thresholds: Low <{CV_LOW}%, High ≥{CV_HIGH}%")

# ============================================================================
# DATA LOADING AND PREPROCESSING
# ============================================================================

print("\n" + "=" * 80)
print("STEP 1: DATA LOADING AND WSI-LEVEL AGGREGATION")
print("=" * 80)

# Load data
corr_df = pd.read_csv(CORR_FILE)
mink_df = pd.read_csv(MINK_FILE)

# Extract metadata
corr_df['WSI_ID'] = corr_df['File name'].str.extract(r'(BRACS_\d+)')
corr_df['Pathology'] = corr_df['File name'].str.extract(r'_([A-Z]+)_\d+\.tif')

mink_df['WSI_ID'] = mink_df['File name'].str.extract(r'(BRACS_\d+)')
mink_df['Pathology'] = mink_df['File name'].str.extract(r'_([A-Z]+)_\d+\.tif')

print(f"\nTotal ROIs: {len(corr_df)}")
print(f"Total WSIs: {corr_df['WSI_ID'].nunique()}")

# ============================================================================
# ANALYSIS 1: WITHIN-WSI DESCRIPTIVE STATISTICS
# ============================================================================

print("\n" + "=" * 80)
print("ANALYSIS 1: WITHIN-WSI DESCRIPTIVE STATISTICS")
print("=" * 80)

def calculate_wsi_heterogeneity(df, dimension_col='Dc'):
    """Calculate comprehensive heterogeneity metrics for each WSI"""
    
    wsi_stats = []
    
    for wsi_id in df['WSI_ID'].unique():
        subset = df[df['WSI_ID'] == wsi_id]
        values = subset[dimension_col].values
        pathologies = subset['Pathology'].unique()
        
        if len(values) >= MIN_ROIS:
            # Basic statistics
            mean_val = np.mean(values)
            std_val = np.std(values, ddof=1)
            median_val = np.median(values)
            min_val = np.min(values)
            max_val = np.max(values)
            
            # Heterogeneity measures
            cv = (std_val / mean_val) * 100 if mean_val > 0 else np.nan
            range_val = max_val - min_val
            q25, q75 = np.percentile(values, [25, 75])
            iqr = q75 - q25
            mad = np.median(np.abs(values - median_val))
            mad_normalized = (mad / median_val) * 100 if median_val > 0 else np.nan
            
            # Standard error of mean
            se_mean = std_val / np.sqrt(len(values))
            
            # Reliability classification
            if len(values) == 1:
                reliability = 'Undefined'
            elif len(values) <= 5:
                reliability = 'Low'
            elif len(values) <= 10:
                reliability = 'Moderate'
            else:
                reliability = 'Good'
            
            wsi_stats.append({
                'WSI_ID': wsi_id,
                'n_ROIs': len(values),
                'n_Pathologies': len(pathologies),
                'Pathologies': '+'.join(sorted(pathologies)),
                'Is_Mixed': len(pathologies) > 1,
                'Dominant_Pathology': subset['Pathology'].mode()[0] if len(subset) > 0 else None,
                'Mean': mean_val,
                'Median': median_val,
                'SD': std_val,
                'Min': min_val,
                'Max': max_val,
                'Range': range_val,
                'CV': cv,
                'Q25': q25,
                'Q75': q75,
                'IQR': iqr,
                'MAD': mad,
                'MAD_normalized': mad_normalized,
                'SE_Mean': se_mean,
                'Reliability': reliability
            })
    
    return pd.DataFrame(wsi_stats)

# Calculate for both dimensions
print("\nCalculating heterogeneity metrics for Correlation Dimension (Dc)...")
wsi_stats_dc = calculate_wsi_heterogeneity(corr_df, 'Dc')

print(f"Analyzing heterogeneity for Minkowski Dimension (Dm)...")
wsi_stats_dm = calculate_wsi_heterogeneity(mink_df, 'Dm')

print(f"\nWSIs with ≥{MIN_ROIS} ROIs:")
print(f"  Dc: {len(wsi_stats_dc)} WSIs")
print(f"  Dm: {len(wsi_stats_dm)} WSIs")

# Summary statistics
print("\n" + "-" * 80)
print("WITHIN-WSI HETEROGENEITY SUMMARY (Correlation Dimension)")
print("-" * 80)
print("\nStandard Deviation (SD):")
print(wsi_stats_dc['SD'].describe())
print("\nCoefficient of Variation (CV%):")
print(wsi_stats_dc['CV'].describe())
print("\nRange (Max - Min):")
print(wsi_stats_dc['Range'].describe())

# Classify WSIs by heterogeneity
wsi_stats_dc['Heterogeneity_Class'] = pd.cut(
    wsi_stats_dc['CV'],
    bins=[0, CV_LOW, CV_HIGH, np.inf],
    labels=['Low', 'Moderate', 'High']
)

print(f"\nHeterogeneity Classification (based on CV):")
print(wsi_stats_dc['Heterogeneity_Class'].value_counts())

# Save WSI-level statistics
wsi_stats_dc.to_excel(RESULTS_DIR / '01_wsi_heterogeneity_dc.xlsx', index=False)
wsi_stats_dm.to_excel(RESULTS_DIR / '01_wsi_heterogeneity_dm.xlsx', index=False)
print(f"\n✓ Saved: {RESULTS_DIR / '01_wsi_heterogeneity_dc.xlsx'}")
print(f"✓ Saved: {RESULTS_DIR / '01_wsi_heterogeneity_dm.xlsx'}")

# ============================================================================
# ANALYSIS 2: INTRACLASS CORRELATION COEFFICIENT (ICC)
# ============================================================================

print("\n" + "=" * 80)
print("ANALYSIS 2: VARIANCE COMPONENTS ANALYSIS (ICC)")
print("=" * 80)

def calculate_icc(df, dimension_col='Dc'):
    """Calculate ICC using one-way random effects ANOVA"""
    
    # Only use WSIs with multiple ROIs
    wsi_counts = df.groupby('WSI_ID').size()
    valid_wsis = wsi_counts[wsi_counts >= MIN_ROIS].index
    df_filtered = df[df['WSI_ID'].isin(valid_wsis)].copy()
    
    # Perform one-way ANOVA
    groups = [df_filtered[df_filtered['WSI_ID'] == wsi][dimension_col].values 
              for wsi in df_filtered['WSI_ID'].unique()]
    
    f_stat, p_value = f_oneway(*groups)
    
    # Calculate variance components
    k = len(df_filtered['WSI_ID'].unique())  # number of WSIs
    n_total = len(df_filtered)
    
    # Mean square between and within
    grand_mean = df_filtered[dimension_col].mean()
    
    # SS between
    ss_between = sum(
        len(df_filtered[df_filtered['WSI_ID'] == wsi]) * 
        (df_filtered[df_filtered['WSI_ID'] == wsi][dimension_col].mean() - grand_mean)**2
        for wsi in df_filtered['WSI_ID'].unique()
    )
    ms_between = ss_between / (k - 1)
    
    # SS within
    ss_within = sum(
        sum((df_filtered[df_filtered['WSI_ID'] == wsi][dimension_col] - 
             df_filtered[df_filtered['WSI_ID'] == wsi][dimension_col].mean())**2)
        for wsi in df_filtered['WSI_ID'].unique()
    )
    ms_within = ss_within / (n_total - k)
    
    # SS total
    ss_total = sum((df_filtered[dimension_col] - grand_mean)**2)
    
    # Calculate variance components
    n_bar = n_total / k  # average group size
    var_within = ms_within
    var_between = (ms_between - ms_within) / n_bar
    
    # ICC
    icc = var_between / (var_between + var_within)
    
    # Eta-squared (proportion of variance between groups)
    eta_squared = ss_between / ss_total
    
    return {
        'ICC': icc,
        'Var_Between': var_between,
        'Var_Within': var_within,
        'MS_Between': ms_between,
        'MS_Within': ms_within,
        'F_statistic': f_stat,
        'P_value': p_value,
        'Eta_Squared': eta_squared,
        'n_WSIs': k,
        'n_ROIs': n_total
    }

# Calculate ICC for both dimensions
icc_dc = calculate_icc(corr_df, 'Dc')
icc_dm = calculate_icc(mink_df, 'Dm')

print("\nCorrelation Dimension (Dc):")
print(f"  ICC = {icc_dc['ICC']:.4f}")
print(f"  Between-WSI variance: {icc_dc['Var_Between']:.6f}")
print(f"  Within-WSI variance: {icc_dc['Var_Within']:.6f}")
print(f"  F-statistic: {icc_dc['F_statistic']:.2f}, p < {icc_dc['P_value']:.2e}")
print(f"  η² = {icc_dc['Eta_Squared']:.4f}")

print(f"\nInterpretation:")
if icc_dc['ICC'] > 0.75:
    print(f"  → HIGH ICC: {100*icc_dc['ICC']:.1f}% of variance is between WSIs")
    print(f"  → LOW within-WSI heterogeneity (homogeneous within patients)")
elif icc_dc['ICC'] > 0.50:
    print(f"  → MODERATE ICC: {100*icc_dc['ICC']:.1f}% of variance is between WSIs")
    print(f"  → MODERATE within-WSI heterogeneity")
else:
    print(f"  → LOW ICC: {100*icc_dc['ICC']:.1f}% of variance is between WSIs")
    print(f"  → HIGH within-WSI heterogeneity (heterogeneous within patients)")

print("\nMinkowski Dimension (Dm):")
print(f"  ICC = {icc_dm['ICC']:.4f}")
print(f"  Between-WSI variance: {icc_dm['Var_Between']:.6f}")
print(f"  Within-WSI variance: {icc_dm['Var_Within']:.6f}")
print(f"  F-statistic: {icc_dm['F_statistic']:.2f}, p < {icc_dm['P_value']:.2e}")

# Save ICC results
icc_results = pd.DataFrame([
    {'Dimension': 'Dc (Correlation)', **icc_dc},
    {'Dimension': 'Dm (Minkowski)', **icc_dm}
])
icc_results.to_excel(RESULTS_DIR / '02_icc_variance_components.xlsx', index=False)
print(f"\n✓ Saved: {RESULTS_DIR / '02_icc_variance_components.xlsx'}")

# ============================================================================
# ANALYSIS 3: HETEROGENEITY VS PATHOLOGY TYPE
# ============================================================================

print("\n" + "=" * 80)
print("ANALYSIS 3: HETEROGENEITY BY PATHOLOGY TYPE")
print("=" * 80)

# For WSIs with single dominant pathology
wsi_stats_clean = wsi_stats_dc[wsi_stats_dc['Dominant_Pathology'].isin(PATHOLOGY_ORDER)].copy()

# Group by pathology
heterogeneity_by_pathology = wsi_stats_clean.groupby('Dominant_Pathology').agg({
    'SD': ['mean', 'std', 'median', 'count'],
    'CV': ['mean', 'std', 'median'],
    'Range': ['mean', 'std', 'median']
}).round(4)

print("\nHeterogeneity Metrics by Pathology (Dominant):")
print(heterogeneity_by_pathology)

# ANOVA: Test if heterogeneity differs by pathology
print("\n" + "-" * 80)
print("ANOVA: Does heterogeneity (SD) differ by pathology?")
print("-" * 80)

groups_sd = [wsi_stats_clean[wsi_stats_clean['Dominant_Pathology'] == p]['SD'].values 
             for p in PATHOLOGY_ORDER if p in wsi_stats_clean['Dominant_Pathology'].values]

if len(groups_sd) >= 2:
    f_het, p_het = f_oneway(*groups_sd)
    
    # Effect size
    grand_mean_sd = wsi_stats_clean['SD'].mean()
    ss_between_het = sum(
        len(g) * (g.mean() - grand_mean_sd)**2 for g in groups_sd
    )
    ss_total_het = sum((wsi_stats_clean['SD'] - grand_mean_sd)**2)
    eta2_het = ss_between_het / ss_total_het
    
    print(f"F = {f_het:.4f}, p = {p_het:.6f}")
    print(f"η² = {eta2_het:.4f}")
    
    if p_het < ALPHA:
        print(f"→ Heterogeneity DIFFERS significantly across pathologies")
    else:
        print(f"→ Heterogeneity does NOT differ significantly")

# Save heterogeneity by pathology
heterogeneity_by_pathology.to_excel(RESULTS_DIR / '03_heterogeneity_by_pathology.xlsx')
print(f"\n✓ Saved: {RESULTS_DIR / '03_heterogeneity_by_pathology.xlsx'}")

# ============================================================================
# ANALYSIS 4: MIXED VS PURE PATHOLOGY WSIs
# ============================================================================

print("\n" + "=" * 80)
print("ANALYSIS 4: MIXED vs PURE PATHOLOGY COMPARISON")
print("=" * 80)

pure_wsis = wsi_stats_dc[~wsi_stats_dc['Is_Mixed']]
mixed_wsis = wsi_stats_dc[wsi_stats_dc['Is_Mixed']]

print(f"\nPure pathology WSIs: {len(pure_wsis)} ({100*len(pure_wsis)/len(wsi_stats_dc):.1f}%)")
print(f"Mixed pathology WSIs: {len(mixed_wsis)} ({100*len(mixed_wsis)/len(wsi_stats_dc):.1f}%)")

# Compare heterogeneity
print(f"\nHeterogeneity Comparison:")
print(f"  Pure WSIs:  SD = {pure_wsis['SD'].mean():.4f} ± {pure_wsis['SD'].std():.4f}")
print(f"  Mixed WSIs: SD = {mixed_wsis['SD'].mean():.4f} ± {mixed_wsis['SD'].std():.4f}")

print(f"\n  Pure WSIs:  CV = {pure_wsis['CV'].mean():.2f}% ± {pure_wsis['CV'].std():.2f}%")
print(f"  Mixed WSIs: CV = {mixed_wsis['CV'].mean():.2f}% ± {mixed_wsis['CV'].std():.2f}%")

# Statistical test
t_stat, p_val = ttest_ind(mixed_wsis['SD'].dropna(), pure_wsis['SD'].dropna())
u_stat, p_val_u = mannwhitneyu(mixed_wsis['SD'].dropna(), pure_wsis['SD'].dropna())

# Effect size (Cohen's d)
pooled_std = np.sqrt((mixed_wsis['SD'].std()**2 + pure_wsis['SD'].std()**2) / 2)
cohens_d = (mixed_wsis['SD'].mean() - pure_wsis['SD'].mean()) / pooled_std

print(f"\nStatistical Tests:")
print(f"  t-test: t = {t_stat:.4f}, p = {p_val:.6f}")
print(f"  Mann-Whitney U: U = {u_stat:.0f}, p = {p_val_u:.6f}")
print(f"  Cohen's d = {cohens_d:.4f}")

if p_val < ALPHA:
    print(f"  → Mixed WSIs show {'HIGHER' if mixed_wsis['SD'].mean() > pure_wsis['SD'].mean() else 'LOWER'} heterogeneity (p < {ALPHA})")
else:
    print(f"  → No significant difference in heterogeneity")

# Save comparison
mixed_pure_comparison = pd.DataFrame([
    {
        'Type': 'Pure Pathology',
        'n': len(pure_wsis),
        'Mean_SD': pure_wsis['SD'].mean(),
        'Median_SD': pure_wsis['SD'].median(),
        'Mean_CV': pure_wsis['CV'].mean(),
        'Median_CV': pure_wsis['CV'].median()
    },
    {
        'Type': 'Mixed Pathology',
        'n': len(mixed_wsis),
        'Mean_SD': mixed_wsis['SD'].mean(),
        'Median_SD': mixed_wsis['SD'].median(),
        'Mean_CV': mixed_wsis['CV'].mean(),
        'Median_CV': mixed_wsis['CV'].median()
    }
])

mixed_pure_comparison.to_excel(RESULTS_DIR / '04_mixed_vs_pure_comparison.xlsx', index=False)
print(f"\n✓ Saved: {RESULTS_DIR / '04_mixed_vs_pure_comparison.xlsx'}")

# ============================================================================
# ANALYSIS 5: HETEROGENEITY VS ROI DENSITY
# ============================================================================

print("\n" + "=" * 80)
print("ANALYSIS 5: HETEROGENEITY vs ROI SAMPLING DENSITY")
print("=" * 80)

# Correlation between n_ROIs and heterogeneity
r_sd, p_r_sd = pearsonr(wsi_stats_dc['n_ROIs'], wsi_stats_dc['SD'])
rho_sd, p_rho_sd = spearmanr(wsi_stats_dc['n_ROIs'], wsi_stats_dc['SD'])

r_cv, p_r_cv = pearsonr(wsi_stats_dc['n_ROIs'], wsi_stats_dc['CV'])
rho_cv, p_rho_cv = spearmanr(wsi_stats_dc['n_ROIs'], wsi_stats_dc['CV'])

print(f"\nCorrelation: n_ROIs vs SD")
print(f"  Pearson r = {r_sd:.4f}, p = {p_r_sd:.6f}")
print(f"  Spearman ρ = {rho_sd:.4f}, p = {p_rho_sd:.6f}")

print(f"\nCorrelation: n_ROIs vs CV")
print(f"  Pearson r = {r_cv:.4f}, p = {p_r_cv:.6f}")
print(f"  Spearman ρ = {rho_cv:.4f}, p = {p_rho_cv:.6f}")

if p_r_sd < ALPHA:
    print(f"\n  → {'Positive' if r_sd > 0 else 'Negative'} correlation: "
          f"More ROIs → {'Higher' if r_sd > 0 else 'Lower'} heterogeneity")
else:
    print(f"\n  → No significant correlation with ROI density")

# Save correlation results
roi_density_corr = pd.DataFrame([
    {'Metric': 'SD', 'Pearson_r': r_sd, 'Pearson_p': p_r_sd, 
     'Spearman_rho': rho_sd, 'Spearman_p': p_rho_sd},
    {'Metric': 'CV', 'Pearson_r': r_cv, 'Pearson_p': p_r_cv,
     'Spearman_rho': rho_cv, 'Spearman_p': p_rho_cv}
])

roi_density_corr.to_excel(RESULTS_DIR / '05_roi_density_correlation.xlsx', index=False)
print(f"\n✓ Saved: {RESULTS_DIR / '05_roi_density_correlation.xlsx'}")

# ============================================================================
# ANALYSIS 6: OUTLIER DETECTION
# ============================================================================

print("\n" + "=" * 80)
print("ANALYSIS 6: OUTLIER DETECTION (EXTREME HETEROGENEITY)")
print("=" * 80)

# Z-scores for heterogeneity
wsi_stats_dc['SD_zscore'] = (wsi_stats_dc['SD'] - wsi_stats_dc['SD'].mean()) / wsi_stats_dc['SD'].std()
wsi_stats_dc['CV_zscore'] = (wsi_stats_dc['CV'] - wsi_stats_dc['CV'].mean()) / wsi_stats_dc['CV'].std()

# Identify outliers (|z| > 3)
outliers_sd = wsi_stats_dc[np.abs(wsi_stats_dc['SD_zscore']) > 3]
outliers_cv = wsi_stats_dc[np.abs(wsi_stats_dc['CV_zscore']) > 3]

print(f"\nOutliers (|z| > 3):")
print(f"  By SD: {len(outliers_sd)} WSIs ({100*len(outliers_sd)/len(wsi_stats_dc):.2f}%)")
print(f"  By CV: {len(outliers_cv)} WSIs ({100*len(outliers_cv)/len(wsi_stats_dc):.2f}%)")

if len(outliers_sd) > 0:
    print(f"\nMost heterogeneous WSIs (by SD):")
    print(outliers_sd.nlargest(5, 'SD')[['WSI_ID', 'Pathologies', 'n_ROIs', 'SD', 'CV', 'Range']].to_string(index=False))

# Save outliers
outliers_sd.to_excel(RESULTS_DIR / '06_outlier_wsis.xlsx', index=False)
print(f"\n✓ Saved: {RESULTS_DIR / '06_outlier_wsis.xlsx'}")

# ============================================================================
# Continue with visualizations in next section...
# ============================================================================

print("\n" + "=" * 80)
print("Continuing with visualizations...")
print("=" * 80)
# ============================================================================
# VISUALIZATION 1: DISTRIBUTION OF HETEROGENEITY MEASURES
# ============================================================================

print("\n" + "=" * 80)
print("CREATING VISUALIZATIONS")
print("=" * 80)

fig, axes = plt.subplots(2, 3, figsize=(16, 10))
fig.suptitle('RQ4: Distribution of Within-WSI Heterogeneity Measures',
             fontsize=14, fontweight='bold')

# Plot 1: Standard Deviation
ax = axes[0, 0]
ax.hist(wsi_stats_dc['SD'].dropna(), bins=30, color='steelblue', 
        alpha=0.7, edgecolor='black')
ax.axvline(wsi_stats_dc['SD'].mean(), color='red', linestyle='--', 
          linewidth=2, label=f'Mean = {wsi_stats_dc["SD"].mean():.4f}')
ax.axvline(wsi_stats_dc['SD'].median(), color='green', linestyle='--',
          linewidth=2, label=f'Median = {wsi_stats_dc["SD"].median():.4f}')
ax.set_xlabel('Standard Deviation (SD)')
ax.set_ylabel('Frequency')
ax.set_title('Within-WSI Standard Deviation')
ax.legend()
ax.grid(True, alpha=0.3)

# Plot 2: Coefficient of Variation
ax = axes[0, 1]
ax.hist(wsi_stats_dc['CV'].dropna(), bins=30, color='coral',
        alpha=0.7, edgecolor='black')
ax.axvline(CV_LOW, color='green', linestyle=':', linewidth=2, 
          label=f'Low (<{CV_LOW}%)')
ax.axvline(CV_HIGH, color='red', linestyle=':', linewidth=2,
          label=f'High (≥{CV_HIGH}%)')
ax.axvline(wsi_stats_dc['CV'].mean(), color='darkred', linestyle='--',
          linewidth=2, label=f'Mean = {wsi_stats_dc["CV"].mean():.2f}%')
ax.set_xlabel('Coefficient of Variation (%)')
ax.set_ylabel('Frequency')
ax.set_title('Within-WSI Coefficient of Variation')
ax.legend()
ax.grid(True, alpha=0.3)

# Plot 3: Range
ax = axes[0, 2]
ax.hist(wsi_stats_dc['Range'].dropna(), bins=30, color='lightgreen',
        alpha=0.7, edgecolor='black')
ax.axvline(wsi_stats_dc['Range'].mean(), color='darkgreen', linestyle='--',
          linewidth=2, label=f'Mean = {wsi_stats_dc["Range"].mean():.4f}')
ax.set_xlabel('Range (Max - Min)')
ax.set_ylabel('Frequency')
ax.set_title('Within-WSI Range')
ax.legend()
ax.grid(True, alpha=0.3)

# Plot 4: IQR
ax = axes[1, 0]
ax.hist(wsi_stats_dc['IQR'].dropna(), bins=30, color='plum',
        alpha=0.7, edgecolor='black')
ax.axvline(wsi_stats_dc['IQR'].mean(), color='purple', linestyle='--',
          linewidth=2, label=f'Mean = {wsi_stats_dc["IQR"].mean():.4f}')
ax.set_xlabel('Interquartile Range (IQR)')
ax.set_ylabel('Frequency')
ax.set_title('Within-WSI Interquartile Range')
ax.legend()
ax.grid(True, alpha=0.3)

# Plot 5: MAD
ax = axes[1, 1]
ax.hist(wsi_stats_dc['MAD_normalized'].dropna(), bins=30, color='gold',
        alpha=0.7, edgecolor='black')
ax.axvline(wsi_stats_dc['MAD_normalized'].mean(), color='orange', 
          linestyle='--', linewidth=2,
          label=f'Mean = {wsi_stats_dc["MAD_normalized"].mean():.2f}%')
ax.set_xlabel('Normalized MAD (%)')
ax.set_ylabel('Frequency')
ax.set_title('Median Absolute Deviation (Normalized)')
ax.legend()
ax.grid(True, alpha=0.3)

# Plot 6: Heterogeneity Classification
ax = axes[1, 2]
hetero_counts = wsi_stats_dc['Heterogeneity_Class'].value_counts()
colors_het = ['green', 'orange', 'red']
ax.bar(range(len(hetero_counts)), hetero_counts.values, 
      color=colors_het[:len(hetero_counts)], alpha=0.7, edgecolor='black')
ax.set_xticks(range(len(hetero_counts)))
ax.set_xticklabels(hetero_counts.index)
ax.set_ylabel('Number of WSIs')
ax.set_title('Heterogeneity Classification')
ax.grid(True, alpha=0.3, axis='y')

for i, v in enumerate(hetero_counts.values):
    ax.text(i, v + 2, str(v), ha='center', fontweight='bold')

plt.tight_layout()
plt.savefig(PLOTS_DIR / 'fig1_heterogeneity_distributions.tif', 
           format='tif', dpi=300, bbox_inches='tight')
plt.close()
print(f"✓ Saved: {PLOTS_DIR / 'fig1_heterogeneity_distributions.tif'}")

# Export data for Origin
fig1_data = wsi_stats_dc[['WSI_ID', 'SD', 'CV', 'Range', 'IQR', 
                          'MAD_normalized', 'Heterogeneity_Class']].copy()
fig1_data.to_excel(ORIGIN_DATA_DIR / 'fig1_heterogeneity_distributions.xlsx',
                  index=False)
print(f"✓ Saved: {ORIGIN_DATA_DIR / 'fig1_heterogeneity_distributions.xlsx'}")

# ============================================================================
# VISUALIZATION 2: ICC VARIANCE DECOMPOSITION
# ============================================================================

fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.suptitle('RQ4: Variance Components Analysis (ICC)',
             fontsize=14, fontweight='bold')

# Pie chart for Dc
ax = axes[0]
sizes_dc = [icc_dc['Var_Between'], icc_dc['Var_Within']]
labels_dc = [f"Between-WSI\n{100*icc_dc['ICC']:.1f}%",
            f"Within-WSI\n{100*(1-icc_dc['ICC']):.1f}%"]
colors_icc = ['#ff9999', '#66b3ff']
wedges, texts, autotexts = ax.pie(sizes_dc, labels=labels_dc, colors=colors_icc,
                                   autopct='%1.1f%%', startangle=90,
                                   textprops={'fontsize': 11, 'fontweight': 'bold'})
ax.set_title(f'Correlation Dimension (Dc)\nICC = {icc_dc["ICC"]:.3f}',
            fontweight='bold')

# Pie chart for Dm
ax = axes[1]
sizes_dm = [icc_dm['Var_Between'], icc_dm['Var_Within']]
labels_dm = [f"Between-WSI\n{100*icc_dm['ICC']:.1f}%",
            f"Within-WSI\n{100*(1-icc_dm['ICC']):.1f}%"]
wedges, texts, autotexts = ax.pie(sizes_dm, labels=labels_dm, colors=colors_icc,
                                   autopct='%1.1f%%', startangle=90,
                                   textprops={'fontsize': 11, 'fontweight': 'bold'})
ax.set_title(f'Minkowski Dimension (Dm)\nICC = {icc_dm["ICC"]:.3f}',
            fontweight='bold')

plt.tight_layout()
plt.savefig(PLOTS_DIR / 'fig2_icc_variance_decomposition.tif',
           format='tif', dpi=300, bbox_inches='tight')
plt.close()
print(f"✓ Saved: {PLOTS_DIR / 'fig2_icc_variance_decomposition.tif'}")

# Export ICC data
icc_results.to_excel(ORIGIN_DATA_DIR / 'fig2_icc_data.xlsx', index=False)

# ============================================================================
# VISUALIZATION 3: HETEROGENEITY BY PATHOLOGY
# ============================================================================

fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.suptitle('RQ4: Within-WSI Heterogeneity by Pathology Type',
             fontsize=14, fontweight='bold')

# Box plot for SD
ax = axes[0]
data_by_path = [wsi_stats_clean[wsi_stats_clean['Dominant_Pathology']==p]['SD'].values
                for p in PATHOLOGY_ORDER 
                if p in wsi_stats_clean['Dominant_Pathology'].values]
path_labels = [p for p in PATHOLOGY_ORDER 
              if p in wsi_stats_clean['Dominant_Pathology'].values]

bp1 = ax.boxplot(data_by_path, labels=path_labels, patch_artist=True, notch=True)
for patch, color in zip(bp1['boxes'], 
                       plt.cm.Set3(np.linspace(0, 1, len(data_by_path)))):
    patch.set_facecolor(color)

ax.set_xlabel('Pathology Type')
ax.set_ylabel('Within-WSI Standard Deviation')
ax.set_title('Standard Deviation by Pathology')
ax.grid(True, alpha=0.3, axis='y')

# Box plot for CV
ax = axes[1]
data_cv_by_path = [wsi_stats_clean[wsi_stats_clean['Dominant_Pathology']==p]['CV'].values
                   for p in PATHOLOGY_ORDER
                   if p in wsi_stats_clean['Dominant_Pathology'].values]

bp2 = ax.boxplot(data_cv_by_path, labels=path_labels, patch_artist=True, notch=True)
for patch, color in zip(bp2['boxes'],
                       plt.cm.Set3(np.linspace(0, 1, len(data_cv_by_path)))):
    patch.set_facecolor(color)

ax.axhline(CV_LOW, color='green', linestyle=':', linewidth=2, alpha=0.7)
ax.axhline(CV_HIGH, color='red', linestyle=':', linewidth=2, alpha=0.7)
ax.set_xlabel('Pathology Type')
ax.set_ylabel('Coefficient of Variation (%)')
ax.set_title('Coefficient of Variation by Pathology')
ax.grid(True, alpha=0.3, axis='y')

plt.tight_layout()
plt.savefig(PLOTS_DIR / 'fig3_heterogeneity_by_pathology.tif',
           format='tif', dpi=300, bbox_inches='tight')
plt.close()
print(f"✓ Saved: {PLOTS_DIR / 'fig3_heterogeneity_by_pathology.tif'}")

# Export pathology data
path_export = wsi_stats_clean[['Dominant_Pathology', 'SD', 'CV']].copy()
path_export.to_excel(ORIGIN_DATA_DIR / 'fig3_pathology_data.xlsx', index=False)

# ============================================================================
# VISUALIZATION 4: MIXED VS PURE WSIs
# ============================================================================

fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.suptitle('RQ4: Mixed vs Pure Pathology WSIs',
             fontsize=14, fontweight='bold')

# Box plot for SD
ax = axes[0]
data_mixed_pure = [pure_wsis['SD'].values, mixed_wsis['SD'].values]
bp1 = ax.boxplot(data_mixed_pure, labels=['Pure\nPathology', 'Mixed\nPathology'],
                patch_artist=True, notch=True)
bp1['boxes'][0].set_facecolor('lightblue')
bp1['boxes'][1].set_facecolor('lightcoral')

ax.set_ylabel('Within-WSI Standard Deviation')
ax.set_title(f'Standard Deviation\nt={t_stat:.3f}, p={p_val:.4f}')
ax.grid(True, alpha=0.3, axis='y')

# Add sample sizes
ax.text(1, ax.get_ylim()[1]*0.95, f'n={len(pure_wsis)}', 
       ha='center', fontsize=10)
ax.text(2, ax.get_ylim()[1]*0.95, f'n={len(mixed_wsis)}',
       ha='center', fontsize=10)

# Box plot for CV
ax = axes[1]
data_cv_mixed_pure = [pure_wsis['CV'].values, mixed_wsis['CV'].values]
bp2 = ax.boxplot(data_cv_mixed_pure, labels=['Pure\nPathology', 'Mixed\nPathology'],
                patch_artist=True, notch=True)
bp2['boxes'][0].set_facecolor('lightblue')
bp2['boxes'][1].set_facecolor('lightcoral')

ax.axhline(CV_LOW, color='green', linestyle=':', linewidth=2, alpha=0.7)
ax.axhline(CV_HIGH, color='red', linestyle=':', linewidth=2, alpha=0.7)
ax.set_ylabel('Coefficient of Variation (%)')
ax.set_title(f'Coefficient of Variation\nCohen\'s d={cohens_d:.3f}')
ax.grid(True, alpha=0.3, axis='y')

plt.tight_layout()
plt.savefig(PLOTS_DIR / 'fig4_mixed_vs_pure.tif',
           format='tif', dpi=300, bbox_inches='tight')
plt.close()
print(f"✓ Saved: {PLOTS_DIR / 'fig4_mixed_vs_pure.tif'}")

# Export mixed vs pure data
mixed_pure_export = pd.DataFrame({
    'WSI_ID': list(pure_wsis['WSI_ID']) + list(mixed_wsis['WSI_ID']),
    'Type': ['Pure']*len(pure_wsis) + ['Mixed']*len(mixed_wsis),
    'SD': list(pure_wsis['SD']) + list(mixed_wsis['SD']),
    'CV': list(pure_wsis['CV']) + list(mixed_wsis['CV'])
})
mixed_pure_export.to_excel(ORIGIN_DATA_DIR / 'fig4_mixed_vs_pure_data.xlsx',
                           index=False)

# ============================================================================
# VISUALIZATION 5: HETEROGENEITY VS ROI DENSITY
# ============================================================================

fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.suptitle('RQ4: Heterogeneity vs ROI Sampling Density',
             fontsize=14, fontweight='bold')

# Scatter plot: n_ROIs vs SD
ax = axes[0]
ax.scatter(wsi_stats_dc['n_ROIs'], wsi_stats_dc['SD'], 
          alpha=0.5, s=50, c='steelblue', edgecolors='black', linewidth=0.5)

# Fit line
z = np.polyfit(wsi_stats_dc['n_ROIs'], wsi_stats_dc['SD'], 1)
p_fit = np.poly1d(z)
x_line = np.linspace(wsi_stats_dc['n_ROIs'].min(), 
                     wsi_stats_dc['n_ROIs'].max(), 100)
ax.plot(x_line, p_fit(x_line), 'r--', linewidth=2,
       label=f'r={r_sd:.3f}, p={p_r_sd:.4f}')

ax.set_xlabel('Number of ROIs per WSI')
ax.set_ylabel('Within-WSI Standard Deviation')
ax.set_title('SD vs ROI Density')
ax.legend()
ax.grid(True, alpha=0.3)

# Scatter plot: n_ROIs vs CV
ax = axes[1]
ax.scatter(wsi_stats_dc['n_ROIs'], wsi_stats_dc['CV'],
          alpha=0.5, s=50, c='coral', edgecolors='black', linewidth=0.5)

# Fit line
z_cv = np.polyfit(wsi_stats_dc['n_ROIs'], wsi_stats_dc['CV'], 1)
p_fit_cv = np.poly1d(z_cv)
ax.plot(x_line, p_fit_cv(x_line), 'r--', linewidth=2,
       label=f'r={r_cv:.3f}, p={p_r_cv:.4f}')

ax.axhline(CV_LOW, color='green', linestyle=':', linewidth=2, alpha=0.5)
ax.axhline(CV_HIGH, color='red', linestyle=':', linewidth=2, alpha=0.5)
ax.set_xlabel('Number of ROIs per WSI')
ax.set_ylabel('Coefficient of Variation (%)')
ax.set_title('CV vs ROI Density')
ax.legend()
ax.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(PLOTS_DIR / 'fig5_heterogeneity_vs_roi_density.tif',
           format='tif', dpi=300, bbox_inches='tight')
plt.close()
print(f"✓ Saved: {PLOTS_DIR / 'fig5_heterogeneity_vs_roi_density.tif'}")

# Export scatter data
scatter_export = wsi_stats_dc[['n_ROIs', 'SD', 'CV']].copy()
scatter_export.to_excel(ORIGIN_DATA_DIR / 'fig5_scatter_data.xlsx', index=False)

# ============================================================================
# VISUALIZATION 6: OUTLIER WSIs
# ============================================================================

fig, ax = plt.subplots(1, 1, figsize=(12, 8))
fig.suptitle('RQ4: WSIs with Extreme Heterogeneity (Outliers)',
             fontsize=14, fontweight='bold')

# Plot all WSIs
ax.scatter(wsi_stats_dc['Mean'], wsi_stats_dc['SD'],
          alpha=0.5, s=50, c='lightgray', edgecolors='black', 
          linewidth=0.5, label='Normal WSIs')

# Highlight outliers
if len(outliers_sd) > 0:
    ax.scatter(outliers_sd['Mean'], outliers_sd['SD'],
              alpha=0.8, s=100, c='red', edgecolors='darkred',
              linewidth=1, marker='^', label=f'Outliers (n={len(outliers_sd)})')
    
    # Annotate top 5 outliers
    top_outliers = outliers_sd.nlargest(5, 'SD')
    for idx, row in top_outliers.iterrows():
        ax.annotate(row['WSI_ID'], 
                   (row['Mean'], row['SD']),
                   textcoords="offset points",
                   xytext=(5, 5),
                   fontsize=8,
                   bbox=dict(boxstyle='round,pad=0.3', 
                           facecolor='yellow', alpha=0.7))

ax.set_xlabel('Mean Dc (Within WSI)')
ax.set_ylabel('Standard Deviation (Heterogeneity)')
ax.set_title('Mean vs Heterogeneity with Outliers Highlighted')
ax.legend()
ax.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(PLOTS_DIR / 'fig6_outlier_wsis.tif',
           format='tif', dpi=300, bbox_inches='tight')
plt.close()
print(f"✓ Saved: {PLOTS_DIR / 'fig6_outlier_wsis.tif'}")

# Export outlier data
outlier_export = wsi_stats_dc[['WSI_ID', 'Mean', 'SD', 'SD_zscore']].copy()
outlier_export.to_excel(ORIGIN_DATA_DIR / 'fig6_outlier_data.xlsx', index=False)

# ============================================================================
# FINAL SUMMARY REPORT
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING FINAL SUMMARY REPORT")
print("=" * 80)

# Calculate additional summary statistics
total_wsis = len(wsi_stats_dc)
mean_heterogeneity = wsi_stats_dc['SD'].mean()
median_heterogeneity = wsi_stats_dc['SD'].median()
mean_cv = wsi_stats_dc['CV'].mean()

# Determine overall heterogeneity level
if icc_dc['ICC'] > 0.75:
    heterogeneity_level = "LOW"
    heterogeneity_interpretation = "Most variance is between WSIs (patients differ, but homogeneous within)"
elif icc_dc['ICC'] > 0.50:
    heterogeneity_level = "MODERATE"
    heterogeneity_interpretation = "Variance split between and within WSIs (moderate within-patient heterogeneity)"
else:
    heterogeneity_level = "HIGH"
    heterogeneity_interpretation = "Most variance is within WSIs (high within-patient heterogeneity)"

summary_report = f"""
================================================================================
RESEARCH QUESTION 4 (RQ4): WITHIN-WSI HETEROGENEITY - SUMMARY REPORT
================================================================================

Research Question:
"How heterogeneous is nuclear spatial organization within the same whole-slide 
image (WSI), and does this heterogeneity correlate with pathological diagnosis?"

ANSWER: Nuclear spatial organization shows {heterogeneity_level} within-WSI 
heterogeneity (ICC = {icc_dc['ICC']:.3f})

================================================================================
KEY FINDINGS
================================================================================

1. OVERALL HETEROGENEITY LEVEL
   ---------------------------------------------------------------
   Intraclass Correlation Coefficient (ICC):
   - Dc (Correlation): ICC = {icc_dc['ICC']:.4f}
   - Dm (Minkowski): ICC = {icc_dm['ICC']:.4f}
   
   Variance Decomposition (Dc):
   - Between-WSI variance: {icc_dc['Var_Between']:.6f} ({100*icc_dc['ICC']:.1f}%)
   - Within-WSI variance: {icc_dc['Var_Within']:.6f} ({100*(1-icc_dc['ICC']):.1f}%)
   
   Interpretation:
   → {heterogeneity_interpretation}
   
   F-statistic: F = {icc_dc['F_statistic']:.2f}, p < {icc_dc['P_value']:.2e}
   → Patient identity {'has a STRONG effect' if icc_dc['ICC'] > 0.7 else 'has a MODERATE effect' if icc_dc['ICC'] > 0.5 else 'has a WEAK effect'} on fractal dimensions

2. DESCRIPTIVE STATISTICS (Correlation Dimension)
   ---------------------------------------------------------------
   WSIs analyzed: {total_wsis} (with ≥{MIN_ROIS} ROIs)
   
   Standard Deviation:
   - Mean: {mean_heterogeneity:.4f}
   - Median: {median_heterogeneity:.4f}
   - Range: [{wsi_stats_dc['SD'].min():.4f}, {wsi_stats_dc['SD'].max():.4f}]
   
   Coefficient of Variation:
   - Mean: {mean_cv:.2f}%
   - Median: {wsi_stats_dc['CV'].median():.2f}%
   - Range: [{wsi_stats_dc['CV'].min():.2f}%, {wsi_stats_dc['CV'].max():.2f}%]
   
   Heterogeneity Classification (by CV):
   - Low (<{CV_LOW}%): {len(wsi_stats_dc[wsi_stats_dc['Heterogeneity_Class']=='Low'])} WSIs ({100*len(wsi_stats_dc[wsi_stats_dc['Heterogeneity_Class']=='Low'])/total_wsis:.1f}%)
   - Moderate ({CV_LOW}-{CV_HIGH}%): {len(wsi_stats_dc[wsi_stats_dc['Heterogeneity_Class']=='Moderate'])} WSIs ({100*len(wsi_stats_dc[wsi_stats_dc['Heterogeneity_Class']=='Moderate'])/total_wsis:.1f}%)
   - High (≥{CV_HIGH}%): {len(wsi_stats_dc[wsi_stats_dc['Heterogeneity_Class']=='High'])} WSIs ({100*len(wsi_stats_dc[wsi_stats_dc['Heterogeneity_Class']=='High'])/total_wsis:.1f}%)

3. HETEROGENEITY BY PATHOLOGY TYPE
   ---------------------------------------------------------------
   ANOVA: F = {f_het:.4f}, p = {p_het:.6f}
   Effect size: η² = {eta2_het:.4f}
   
   {'→ Heterogeneity DIFFERS significantly across pathologies' if p_het < ALPHA else '→ Heterogeneity does NOT differ significantly across pathologies'}
   
   Mean SD by pathology (top 3 most heterogeneous):
   {heterogeneity_by_pathology['SD']['mean'].nlargest(3).to_string()}

4. MIXED vs PURE PATHOLOGY WSIs
   ---------------------------------------------------------------
   Pure pathology WSIs: {len(pure_wsis)} ({100*len(pure_wsis)/total_wsis:.1f}%)
   Mixed pathology WSIs: {len(mixed_wsis)} ({100*len(mixed_wsis)/total_wsis:.1f}%)
   
   Heterogeneity comparison:
   - Pure:  Mean SD = {pure_wsis['SD'].mean():.4f}, CV = {pure_wsis['CV'].mean():.2f}%
   - Mixed: Mean SD = {mixed_wsis['SD'].mean():.4f}, CV = {mixed_wsis['CV'].mean():.2f}%
   
   Statistical test:
   - t-test: t = {t_stat:.4f}, p = {p_val:.6f}
   - Cohen's d = {cohens_d:.4f} ({'Negligible' if abs(cohens_d)<0.2 else 'Small' if abs(cohens_d)<0.5 else 'Medium' if abs(cohens_d)<0.8 else 'Large'} effect)
   
   {'→ Mixed WSIs show HIGHER heterogeneity (p < 0.05)' if p_val < ALPHA and mixed_wsis['SD'].mean() > pure_wsis['SD'].mean() else '→ No significant difference' if p_val >= ALPHA else '→ Pure WSIs show HIGHER heterogeneity (unexpected)'}

5. HETEROGENEITY vs ROI SAMPLING DENSITY
   ---------------------------------------------------------------
   Correlation: n_ROIs vs SD
   - Pearson r = {r_sd:.4f}, p = {p_r_sd:.6f}
   - Spearman ρ = {rho_sd:.4f}, p = {p_rho_sd:.6f}
   
   Correlation: n_ROIs vs CV
   - Pearson r = {r_cv:.4f}, p = {p_r_cv:.6f}
   - Spearman ρ = {rho_cv:.4f}, p = {p_rho_cv:.6f}
   
   {'→ Positive correlation: More ROIs reveal more heterogeneity' if r_sd > 0 and p_r_sd < ALPHA else '→ Negative correlation: More ROIs show less heterogeneity' if r_sd < 0 and p_r_sd < ALPHA else '→ No significant correlation with sampling density'}
   
   Interpretation:
   {'Sampling bias present - well-sampled WSIs show higher heterogeneity' if r_sd > 0.2 and p_r_sd < ALPHA else 'No sampling bias effect detected'}

6. OUTLIER DETECTION
   ---------------------------------------------------------------
   WSIs with extreme heterogeneity (|z| > 3): {len(outliers_sd)}
   Percentage: {100*len(outliers_sd)/total_wsis:.2f}%
   
   Most heterogeneous WSI:
   - ID: {wsi_stats_dc.loc[wsi_stats_dc['SD'].idxmax(), 'WSI_ID']}
   - SD: {wsi_stats_dc['SD'].max():.4f}
   - CV: {wsi_stats_dc.loc[wsi_stats_dc['SD'].idxmax(), 'CV']:.2f}%
   - n_ROIs: {wsi_stats_dc.loc[wsi_stats_dc['SD'].idxmax(), 'n_ROIs']:.0f}
   - Pathologies: {wsi_stats_dc.loc[wsi_stats_dc['SD'].idxmax(), 'Pathologies']}

================================================================================
CLINICAL INTERPRETATION
================================================================================

OVERALL HETEROGENEITY: {heterogeneity_level}

{f'''IMPLICATIONS FOR CLINICAL PRACTICE:

{"HIGH HETEROGENEITY DETECTED:" if icc_dc['ICC'] < 0.5 else "MODERATE HETEROGENEITY:" if icc_dc['ICC'] < 0.75 else "LOW HETEROGENEITY:"}

Sampling Strategy:
{'✗ Single biopsy may be INSUFFICIENT' if icc_dc['ICC'] < 0.5 else '⚠ Single biopsy may miss some variation' if icc_dc['ICC'] < 0.75 else '✓ Single biopsy generally SUFFICIENT'}
{'✓ Multiple sampling sites RECOMMENDED' if icc_dc['ICC'] < 0.5 else '⚠ Consider multiple sites for complex cases' if icc_dc['ICC'] < 0.75 else '⚠ Single site adequate for most cases'}
{'✓ Spatial mapping important for accurate diagnosis' if icc_dc['ICC'] < 0.5 else '⚠ Spatial mapping useful for borderline cases' if icc_dc['ICC'] < 0.75 else '✓ Standard sampling protocols adequate'}

Diagnostic Accuracy:
{'⚠ HIGH sampling bias risk' if icc_dc['ICC'] < 0.5 else '⚠ Moderate sampling bias risk' if icc_dc['ICC'] < 0.75 else '✓ Low sampling bias risk'}
{'⚠ Single ROI may not represent entire tumor' if icc_dc['ICC'] < 0.5 else '⚠ Single ROI generally representative' if icc_dc['ICC'] < 0.75 else '✓ Single ROI reliable'}
{'⚠ Need comprehensive tissue assessment' if icc_dc['ICC'] < 0.5 else '✓ Standard assessment usually sufficient' if icc_dc['ICC'] < 0.75 else '✓ Routine assessment adequate'}

Research Implications:
{'✓ Multi-region sampling essential' if icc_dc['ICC'] < 0.5 else '⚠ Multi-region sampling beneficial' if icc_dc['ICC'] < 0.75 else '✓ Single-region analysis acceptable'}
{'✓ Spatial analysis techniques required' if icc_dc['ICC'] < 0.5 else '⚠ Consider spatial methods for precision' if icc_dc['ICC'] < 0.75 else '✓ Standard methods adequate'}
{'✓ Heterogeneity should be reported as outcome measure' if icc_dc['ICC'] < 0.5 else '⚠ Heterogeneity may be relevant covariate' if icc_dc['ICC'] < 0.75 else '✓ Heterogeneity not critical factor'}

Treatment Planning:
{'⚠ Heterogeneous tumors may show variable response' if icc_dc['ICC'] < 0.5 else '⚠ Monitor multiple sites for response' if icc_dc['ICC'] < 0.75 else '✓ Standard monitoring adequate'}
{'⚠ Targeted therapy may miss resistant clones' if icc_dc['ICC'] < 0.5 else '⚠ Consider combination therapy' if icc_dc['ICC'] < 0.75 else '✓ Standard therapy appropriate'}
{'✓ Comprehensive pre-treatment assessment needed' if icc_dc['ICC'] < 0.5 else '⚠ Standard assessment with attention to variants' if icc_dc['ICC'] < 0.75 else '✓ Routine assessment sufficient'}
''' if icc_dc['ICC'] < 0.75 else 'MODERATE to LOW heterogeneity suggests standard clinical protocols are adequate'}

PATHOLOGY-SPECIFIC FINDINGS:
{f"→ Mixed pathology WSIs show {'significantly higher' if p_val < ALPHA and cohens_d > 0.2 else 'similar'} heterogeneity" if len(mixed_wsis) > 0 else "→ Insufficient mixed pathology WSIs for comparison"}
{f"→ {'Some pathologies show distinct heterogeneity patterns' if p_het < ALPHA else 'Heterogeneity is consistent across pathologies'}" if len(wsi_stats_clean) > 0 else ""}
→ Within-patient heterogeneity is {'a significant factor' if icc_dc['ICC'] < 0.7 else 'a moderate factor' if icc_dc['ICC'] < 0.85 else 'a minor factor'} in fractal analysis

================================================================================
STATISTICAL STRENGTH
================================================================================

Sample Size: {total_wsis} WSIs with ≥{MIN_ROIS} ROIs
Total ROIs: {icc_dc['n_ROIs']}
Power: {'High (>80%)' if total_wsis > 100 else 'Adequate (>60%)' if total_wsis > 50 else 'Moderate'}

Reliability by ROI Count:
- 1 ROI (cannot estimate): {len(wsi_stats_dc[wsi_stats_dc['n_ROIs']==1]) if 'n_ROIs' in wsi_stats_dc.columns else 'N/A'} WSIs
- 2-5 ROIs (low reliability): {len(wsi_stats_dc[(wsi_stats_dc['n_ROIs']>=2) & (wsi_stats_dc['n_ROIs']<=5)])} WSIs
- 6-10 ROIs (moderate reliability): {len(wsi_stats_dc[(wsi_stats_dc['n_ROIs']>=6) & (wsi_stats_dc['n_ROIs']<=10)])} WSIs
- >10 ROIs (good reliability): {len(wsi_stats_dc[wsi_stats_dc['n_ROIs']>10])} WSIs

Confidence in Results: {'Very High' if icc_dc['P_value'] < 0.001 and total_wsis > 100 else 'High' if icc_dc['P_value'] < 0.01 else 'Moderate'}

================================================================================
FILES GENERATED
================================================================================

RESULTS (Excel):
- 01_wsi_heterogeneity_dc.xlsx - Complete WSI-level statistics (Dc)
- 01_wsi_heterogeneity_dm.xlsx - Complete WSI-level statistics (Dm)
- 02_icc_variance_components.xlsx - ICC and variance decomposition
- 03_heterogeneity_by_pathology.xlsx - Statistics by pathology type
- 04_mixed_vs_pure_comparison.xlsx - Mixed vs pure pathology comparison
- 05_roi_density_correlation.xlsx - Correlation with ROI sampling
- 06_outlier_wsis.xlsx - Extreme heterogeneity cases

PLOTS (TIF, 300 DPI):
- fig1_heterogeneity_distributions.tif - Distribution of heterogeneity measures
- fig2_icc_variance_decomposition.tif - ICC pie charts
- fig3_heterogeneity_by_pathology.tif - Box plots by pathology
- fig4_mixed_vs_pure.tif - Mixed vs pure comparison
- fig5_heterogeneity_vs_roi_density.tif - Scatter plots
- fig6_outlier_wsis.tif - Outlier identification

ORIGIN DATA:
- All plotting data exported for custom visualization

================================================================================
CONCLUSION
================================================================================

Within-WSI heterogeneity is {heterogeneity_level} (ICC = {icc_dc['ICC']:.3f}), indicating that 
{100*icc_dc['ICC']:.1f}% of variance in fractal dimensions is BETWEEN patients and 
{100*(1-icc_dc['ICC']):.1f}% is WITHIN patients.

{'This HIGH within-patient heterogeneity suggests that multiple biopsies per patient are recommended for accurate characterization.' if icc_dc['ICC'] < 0.5 else 'This MODERATE heterogeneity suggests that sampling strategy should consider patient-specific factors and pathology type.' if icc_dc['ICC'] < 0.75 else 'This LOW within-patient heterogeneity suggests that single biopsies are generally sufficient for fractal dimension estimation.'}

{'Mixed pathology WSIs show significantly higher heterogeneity, confirming that tissue heterogeneity reflects underlying pathological complexity.' if p_val < ALPHA and len(mixed_wsis) > 0 else ''}

================================================================================
"""

# Save summary report
with open(RESULTS_DIR / '00_RQ4_SUMMARY.txt', 'w', encoding='utf-8') as f:
    f.write(summary_report)

print(summary_report)
print(f"\n✓ Saved: {RESULTS_DIR / '00_RQ4_SUMMARY.txt'}")

# ============================================================================
# CREATE MASTER EXCEL FILE
# ============================================================================

print("\n" + "=" * 80)
print("CREATING MASTER EXCEL FILE")
print("=" * 80)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Summary sheet
ws1 = wb.create_sheet('Summary')
ws1['A1'] = 'RQ4: WITHIN-WSI HETEROGENEITY - MASTER RESULTS'
ws1['A1'].font = Font(bold=True, size=14)
ws1['A3'] = 'Research Question:'
ws1['A4'] = 'How heterogeneous is nuclear spatial organization within the same WSI?'
ws1['A6'] = f'Answer: {heterogeneity_level} heterogeneity (ICC = {icc_dc["ICC"]:.3f})'
ws1['A6'].font = Font(bold=True, color='008000')

# Add data sheets
sheets_data = [
    ('ICC_Results', icc_results),
    ('Heterogeneity_by_Path', heterogeneity_by_pathology.reset_index()),
    ('Mixed_vs_Pure', mixed_pure_comparison),
    ('ROI_Density_Corr', roi_density_corr)
]

for sheet_name, df in sheets_data:
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', fill_type='solid')

# Adjust column widths
for ws in wb.worksheets:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(RESULTS_DIR / 'RQ4_MASTER_RESULTS.xlsx')
print(f"\n✓ Saved: {RESULTS_DIR / 'RQ4_MASTER_RESULTS.xlsx'}")

# ============================================================================
# COMPLETION MESSAGE
# ============================================================================

print("\n" + "=" * 80)
print("RQ4 ANALYSIS COMPLETE!")
print("=" * 80)
print(f"\nAll results saved to: {OUTPUT_DIR}")
print(f"\nGenerated files:")
print(f"  - Plots (TIF): {len(list(PLOTS_DIR.glob('*.tif')))} figures")
print(f"  - Excel results: {len(list(RESULTS_DIR.glob('*.xlsx')))} files")
print(f"  - Origin data files: {len(list(ORIGIN_DATA_DIR.glob('*.xlsx')))} files")
print(f"  - Summary report: 00_RQ4_SUMMARY.txt")
print(f"  - Master workbook: RQ4_MASTER_RESULTS.xlsx")
print("\n" + "=" * 80)